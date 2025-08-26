# -*- coding: utf-8 -*-
# 📦 Shopee Mass Upload Excel作成アプリ（フル版）
# - media_info: Option 1〜30 Name/Image を縦持ち化
# - product_id × variation_name（正規化）で突合
# - 公式テンプレの「Image per Variation」列へURLを書き込み
# - 商品説明等を結合
# - 未マッチ一覧CSV / media候補カタログCSVも出力
# ※ ヘッダー破損防止のため、Excelは「2行目から」書き込み

import streamlit as st
import pandas as pd
import numpy as np
import re
import unicodedata
from io import BytesIO
from openpyxl import load_workbook

st.set_page_config(page_title="Shopee Mass Upload Builder", layout="wide")
st.title("📦 Shopee Mass Upload Excel作成アプリ（フル版）")

# 🟡 注意コメント + 補助画像
st.markdown("### ⚠️ STEP1~4に必要なExcelシートは、ダウンロード後に保護解除→**保存し直してから**アップロードしてください。")
st.image("images/unlock_tip.png", width=600)

# === アップローダ ===
basic_info_path    = st.file_uploader("STEP1: basic_info", type=["xlsx"], key="basic")
sales_info_path    = st.file_uploader("STEP2: sales_info", type=["xlsx"], key="sales")
media_info_path    = st.file_uploader("STEP3: media_info", type=["xlsx"], key="media")
shipment_info_path = st.file_uploader("STEP4: shipment_info", type=["xlsx"], key="shipment")
template_path      = st.file_uploader("STEP5: Shopee公式テンプレート", type=["xlsx"], key="template")

# === ユーティリティ ===
INSTRUCTION_ROWS = 5  # mass_update_* の上部説明行の想定行数（必要なら調整）

def normalize_columns(cols):
    """Shopeeテンプレ特有の |n|n サフィックスを除去して処理用名にする"""
    return [re.sub(r"\|\d+\|\d+$", "", str(c)).strip() for c in cols]

def clean_name(x):
    """バリエーション名の正規化（NFKC / 全角空白→半角 / 連続空白→単一 / 改行タブ除去 / strip）"""
    if pd.isna(x):
        return None
    s = unicodedata.normalize("NFKC", str(x))
    s = s.replace("\u3000", " ")
    s = re.sub(r"\s+", " ", s).replace("\r", " ").replace("\n", " ").replace("\t", " ")
    return s.strip() or None

def to_str_id(x):
    """Excelで数値扱いされたID(例: '123456.0')を安全に文字列へ"""
    if pd.isna(x):
        return None
    s = str(x).strip()
    if re.fullmatch(r"\d+\.0", s):
        s = s[:-2]
    return s

def find_image_per_variation_col(norm_cols):
    """テンプレ内の『Image per Variation』列（正規化名）を自動検出"""
    pats = [r"(?i)image\s*per\s*variation", r"(?i)et_title_image_per_variation"]
    for p in pats:
        for c in norm_cols:
            if re.search(p, c):
                return c
    return None

def detect_pairs_media_columns(media_df):
    """
    media_info の実列名に合わせた Name/Image の30ペアを抽出
    - Name: et_title_option_{1..30}_for_variation_1
    - Image: et_title_option_image_{1..30}_for_variation_1
    """
    pairs = []
    for i in range(1, 30 + 1):
        name_col = f"et_title_option_{i}_for_variation_1"
        img_col  = f"et_title_option_image_{i}_for_variation_1"
        if name_col in media_df.columns and img_col in media_df.columns:
            pairs.append((i, name_col, img_col))
    return pairs

def df_to_csv_bytes(df: pd.DataFrame) -> BytesIO:
    buf = BytesIO()
    df.to_csv(buf, index=False, encoding="utf-8-sig")
    buf.seek(0)
    return buf

# === メイン処理 ===
if all([basic_info_path, sales_info_path, media_info_path, shipment_info_path, template_path]):

    # ===== 取り込み =====
    try:
        basic_df    = pd.read_excel(basic_info_path,    sheet_name="Sheet1")
        sales_df    = pd.read_excel(sales_info_path,    sheet_name="Sheet1")
        media_df    = pd.read_excel(media_info_path,    sheet_name="Sheet1")
        shipment_df = pd.read_excel(shipment_info_path, sheet_name="Sheet1")
        template_df = pd.read_excel(template_path,      sheet_name="Template")
    except Exception as e:
        st.error(f"Excel読み込みでエラー: {e}")
        st.stop()

    # 公式列名（順序保持用）
    original_columns     = template_df.columns
    original_cols_normal = normalize_columns(original_columns)

    # テンプレを処理用に正規化
    template_df_norm = template_df.copy()
    template_df_norm.columns = original_cols_normal

    # 欠けている公式列があれば空で補完
    for col in original_cols_normal:
        if col not in template_df_norm.columns:
            template_df_norm[col] = None
    template_df_norm = template_df_norm[original_cols_normal]

    # ===== sales / shipment 等、開始行以降の値を準備 =====
    start = INSTRUCTION_ROWS  # DataFrameインデックスでの開始（上部説明行の下から）
    n = max(0, len(sales_df) - start)
    sl = slice(start, start + n)

    # 必要列が存在する時のみ代入（存在チェックで壊れにくく）
    def safe_set(col, series):
        if col in template_df_norm.columns:
            template_df_norm.loc[sl, col] = series.iloc[start:].values

    safe_set("et_title_variation_integration_no", sales_df["et_title_product_id"])
    safe_set("et_title_variation_id",             sales_df["et_title_variation_id"])
    safe_set("ps_product_name",                   sales_df["et_title_product_name"])
    safe_set("ps_sku_short",                      sales_df["et_title_variation_sku"])
    safe_set("ps_price",                          sales_df["et_title_variation_price"])
    safe_set("ps_stock",                          sales_df["et_title_variation_stock"])
    safe_set("et_title_option_for_variation_1",   sales_df["et_title_variation_name"])
    if "et_title_variation_1" in template_df_norm.columns:
        template_df_norm.loc[sl, "et_title_variation_1"] = "type"
    safe_set("ps_weight",                         shipment_df["et_title_product_weight"])
    if "channel_id.28057" in template_df_norm.columns:
        template_df_norm.loc[sl, "channel_id.28057"] = "On"

    # 価格換算（必要な場合のみONにする）
    with st.expander("💱 価格換算（SGD→MYR）オプション"):
        do_fx = st.checkbox("SGD→MYRを掛ける（3.4×）", value=True)
        fx_rate = st.number_input("レート", value=3.4, step=0.1, format="%.2f")
    if do_fx and "ps_price" in template_df_norm.columns:
        template_df_norm.loc[sl, "ps_price"] = pd.to_numeric(
            template_df_norm.loc[sl, "ps_price"], errors="coerce"
        ).mul(fx_rate).round(2)

    # ===== media_info: Option 1〜30 Name/Image を縦持ち化 =====
    pairs = detect_pairs_media_columns(media_df)
    media_long_list = []
    for _, name_col, img_col in pairs:
        tmp = media_df[["et_title_product_id", name_col, img_col]].copy()
        tmp.rename(columns={name_col: "variation_name_raw", img_col: "variation_image"}, inplace=True)
        tmp["product_id"]      = tmp["et_title_product_id"].map(to_str_id)
        tmp["variation_name"]  = tmp["variation_name_raw"].map(clean_name)
        tmp["variation_image"] = tmp["variation_image"].astype(str).str.strip()
        tmp = tmp[(tmp["product_id"].notna()) & (tmp["variation_name"].notna()) & (tmp["variation_image"] != "")]
        media_long_list.append(tmp[["product_id", "variation_name", "variation_image"]])

    if media_long_list:
        media_long = pd.concat(media_long_list, ignore_index=True)
        # 数値以外のproduct_idを除外（ヘッダーの取り込み事故などを回避）
        media_long = media_long[media_long["product_id"].str.fullmatch(r"\d+")]
        # 同一キーの重複は最初のURLを採用
        media_long = (media_long
                      .sort_values(["product_id", "variation_name"])
                      .drop_duplicates(["product_id", "variation_name"], keep="first"))
    else:
        media_long = pd.DataFrame(columns=["product_id", "variation_name", "variation_image"])

    # ===== sales側キーを正規化（突合用） =====
    variation_map = sales_df[["et_title_product_id", "et_title_variation_name"]].copy()
    variation_map["product_id"]     = variation_map["et_title_product_id"].map(to_str_id)
    variation_map["variation_name"] = variation_map["et_title_variation_name"].map(clean_name)
    variation_map = variation_map[["product_id", "variation_name"]].dropna().drop_duplicates()

    # 画像URLを合流（98%程度が一致するはず）
    img_map = variation_map.merge(media_long, on=["product_id", "variation_name"], how="left")

    # ===== Template側のキーでマージして『Image per Variation』列に書く =====
    template_df_norm["product_id"]     = template_df_norm["et_title_variation_integration_no"].map(to_str_id)
    template_df_norm["variation_name"] = template_df_norm["et_title_option_for_variation_1"].map(clean_name)

    # 実在の Image per Variation 列を検出
    image_per_var_col = find_image_per_variation_col(original_cols_normal)
    if not image_per_var_col:
        st.error("テンプレートに『Image per Variation』列が見つかりません。テンプレ列名をご確認ください。")
        st.stop()

    # マージして画像URLを付与
    template_df_norm = template_df_norm.merge(
        img_map[["product_id", "variation_name", "variation_image"]],
        on=["product_id", "variation_name"],
        how="left"
    )
    template_df_norm.loc[sl, image_per_var_col] = template_df_norm.loc[sl, "variation_image"].values

    # ===== 商品説明を結合 =====
    desc_df = basic_df[["et_title_product_id", "et_title_product_description"]].copy()
    desc_df.rename(columns={"et_title_product_id": "product_id"}, inplace=True)
    desc_df["product_id"] = desc_df["product_id"].map(to_str_id)
    template_df_norm = template_df_norm.merge(desc_df, on="product_id", how="left")
    if "ps_product_description" in template_df_norm.columns:
        template_df_norm.loc[sl, "ps_product_description"] = template_df_norm.loc[sl, "et_title_product_description"].values

    # 一時列の後始末
    template_df_norm.drop(columns=["variation_image", "et_title_product_description", "product_id", "variation_name"],
                          inplace=True, errors="ignore")

    # 列順を公式に整え、表示名を元に戻す
    template_df_norm = template_df_norm[original_cols_normal]
    template_df_norm.columns = original_columns

    # ===== 未マッチCSV / media候補カタログCSV を作成 =====
    # （デバッグ用に sales の元名なども保持）
    variation_map_dbg = sales_df[["et_title_product_id", "et_title_product_name", "et_title_variation_name"]].copy()
    variation_map_dbg["product_id"] = variation_map_dbg["et_title_product_id"].map(to_str_id)
    variation_map_dbg["variation_name_clean"] = variation_map_dbg["et_title_variation_name"].map(clean_name)
    variation_map_dbg = (variation_map_dbg[["product_id", "et_title_product_name", "et_title_variation_name", "variation_name_clean"]]
                         .dropna(subset=["product_id", "variation_name_clean"])
                         .drop_duplicates()
                         .rename(columns={"et_title_product_name": "product_name",
                                          "et_title_variation_name": "variation_name_sales_raw"}))

    img_map_dbg = variation_map.merge(media_long, on=["product_id", "variation_name"], how="left")
    img_map_dbg["has_url"] = img_map_dbg["variation_image"].notna() & (img_map_dbg["variation_image"] != "")

    media_candidates = (media_long.groupby("product_id")["variation_name"]
                        .agg(lambda s: " | ".join(sorted(set([x for x in s if isinstance(x, str)]))))
                        .reset_index()
                        .rename(columns={"variation_name": "media_candidates"}))

    unmatched = img_map_dbg[~img_map_dbg["has_url"]][["product_id", "variation_name"]].copy()
    unmatched = unmatched.rename(columns={"variation_name": "variation_name_clean"})
    unmatched = unmatched.merge(variation_map_dbg, on=["product_id", "variation_name_clean"], how="left")
    unmatched = unmatched.merge(media_candidates, on="product_id", how="left")
    unmatched = unmatched[["product_id", "product_name", "variation_name_sales_raw", "variation_name_clean", "media_candidates"]]

    # ===== Excelへ安全に書き戻し（2行目から書く：ヘッダー保護） =====
    wb = load_workbook(template_path, data_only=True)
    sh = wb["Template"]

    # pandasの行0がExcelの「2行目」に相当
    # → 説明行も含め、DataFrameの内容をそのまま2行目以降に書き戻す
    for r_idx, row in enumerate(template_df_norm.itertuples(index=False, name=None), start=2):
        for c_idx, val in enumerate(row, start=1):
            if isinstance(val, float) and np.isnan(val):
                val = None
            sh.cell(row=r_idx, column=c_idx, value=val)

    # 出力バッファ
    out_xlsx = BytesIO()
    wb.save(out_xlsx)
    out_xlsx.seek(0)

    # CSVバッファ
    unmatched_csv = df_to_csv_bytes(unmatched)
    catalog_csv   = df_to_csv_bytes(media_long[["product_id", "variation_name", "variation_image"]])

    # ===== 画面出力 =====
    matched_cnt = int(img_map_dbg["has_url"].sum())
    total_keys  = len(img_map_dbg)
    st.success(f"✅ 完了：画像URLマッチ {matched_cnt}/{total_keys}（{matched_cnt/total_keys:.2%}）")
    st.download_button(
        "📥 Excelをダウンロード",
        data=out_xlsx,
        file_name="shopee_mass_upload_output.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    st.download_button("⬇️ 未マッチ一覧CSVをダウンロード", data=unmatched_csv,
                       file_name="unmatched_variations.csv", mime="text/csv")
    st.download_button("⬇️ media候補カタログCSVをダウンロード", data=catalog_csv,
                       file_name="media_variation_catalog.csv", mime="text/csv")

else:
    st.info("上の5つのExcelファイルをすべて選択すると処理を実行します。")
